"""
Fase 5 – API de Finanças Globais.

Endpoints:
  GET  /finance/ledger/{supplier_id}          – Extrato da conta corrente (paginado)
  GET  /finance/ledger/{supplier_id}/extract  – Extrato contabilístico por período (com saldo inicial)
  POST /finance/ledger/entry                  – Lançamento manual
  GET  /finance/aging                         – Aging report (dívidas a fornecedores)
  GET  /finance/reconciliation/discrepancies  – Divergências triple-match
  POST /finance/reconciliation/match/{po_id}  – Forçar triple-match para uma PO
  GET  /finance/profitability                 – Lucro Real Final
  GET  /finance/cash-flow-forecast            – Projeção de saídas de caixa
"""
from typing import Any, Dict, List, Optional

from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from app.config.database import get_db_connection
from app.api.deps import get_accounting_service
from app.services.accounting_match_service import AccountingMatchService
from app.services.invoice_service import InvoiceService
from app.services.oss_service import get_oss_report
from app.services.supplier_health_service import SupplierHealthService

router = APIRouter(prefix="/finance", tags=["finance"])


# ─── Schemas ────────────────────────────────────────────────────────────────


class LedgerEntryIn(BaseModel):
    empresa_id: int
    supplier_id: int
    tipo: str
    valor_credito: float = 0.0
    valor_debito: float = 0.0
    documento_ref: Optional[str] = None
    purchase_order_id: Optional[int] = None
    notas: Optional[str] = None
    data_movimento: Optional[str] = None


class LedgerLine(BaseModel):
    id: int
    empresa_id: Optional[int] = None
    supplier_id: Optional[int] = None
    data_movimento: Optional[str] = None
    tipo: str
    documento_ref: Optional[str] = None
    purchase_order_id: Optional[int] = None
    valor_credito: float
    valor_debito: float
    saldo_acumulado: float
    notas: Optional[str] = None
    created_at: Optional[str] = None
    supplier_nome: Optional[str] = None
    empresa_nome: Optional[str] = None


class LedgerResponse(BaseModel):
    entries: List[LedgerLine]
    total: int
    saldo_atual: float


class AgingRow(BaseModel):
    supplier_id: int
    supplier_nome: str
    empresa_id: Optional[int] = None
    empresa_nome: Optional[str] = None
    a_vencer: float
    vencido_30: float
    vencido_mais_30: float
    total_divida: float


class DiscrepancyRow(BaseModel):
    id: int
    empresa_id: Optional[int] = None
    purchase_order_id: Optional[int] = None
    invoice_ref: Optional[str] = None
    invoice_amount: Optional[float] = None
    po_amount: Optional[float] = None
    bank_movement_id: Optional[int] = None
    bank_amount: Optional[float] = None
    status: str
    discrepancy_amount: Optional[float] = None
    discrepancy_notes: Optional[str] = None
    created_at: Optional[str] = None
    supplier_order_id: Optional[str] = None
    data_ordered: Optional[str] = None
    supplier_nome: Optional[str] = None
    empresa_nome: Optional[str] = None


class DiscrepanciesResponse(BaseModel):
    items: List[DiscrepancyRow]
    total: int


class PaymentSuggestionItem(BaseModel):
    purchase_order_id: int
    empresa_id: Optional[int] = None
    empresa_nome: Optional[str] = None
    supplier_id: Optional[int] = None
    supplier_nome: Optional[str] = None
    total_final: float
    data_ordered: Optional[str] = None
    due_date: Optional[str] = None
    invoice_ref: Optional[str] = None
    status: Optional[str] = None
    prazo_pagamento: Optional[str] = None
    metodo_pagamento: Optional[str] = None


class PaymentHistoricoItem(BaseModel):
    ledger_id: int
    data_movimento: Optional[str] = None
    empresa_id: Optional[int] = None
    empresa_nome: Optional[str] = None
    supplier_id: Optional[int] = None
    supplier_nome: Optional[str] = None
    metodo_pagamento: Optional[str] = None
    purchase_order_id: Optional[int] = None
    documento_ref: Optional[str] = None
    valor: float
    notas: Optional[str] = None


class PaymentsAntecipadoResponse(BaseModel):
    items: List[PaymentSuggestionItem]
    total: int


class PaymentsSugestaoResponse(BaseModel):
    items: List[PaymentSuggestionItem]
    total_valor: float


class ConfirmPaymentItem(BaseModel):
    purchase_order_id: int
    valor: Optional[float] = None


class ConfirmPaymentRequest(BaseModel):
    empresa_id: int
    items: List[ConfirmPaymentItem]
    data_pagamento: Optional[str] = None
    criar_movimento_banco: bool = True


class ConfirmPaymentResponse(BaseModel):
    created_ledger_ids: List[int]
    updated_po_ids: List[int]
    reconciliation_ids: List[int]
    bank_movement_id: Optional[int] = None
    total_pago: float


class ProfitabilityResponse(BaseModel):
    gmv: float
    devolucoes: float
    comissoes: float
    custo_base: float
    portes_reais: float
    impostos_po: float
    custo_total: float
    lucro_real: float
    margem_pct: float


class CashFlowDay(BaseModel):
    data_vencimento: str
    total_saidas: float
    num_pos: int
    pos: List[Dict[str, Any]]


class MatchResult(BaseModel):
    purchase_order_id: int
    po_amount: float
    invoice_amount: Optional[float] = None
    bank_amount: Optional[float] = None
    status: str
    discrepancy_amount: float
    discrepancy_notes: Optional[str] = None


# ─── Endpoints ──────────────────────────────────────────────────────────────


@router.get("/ledger/{supplier_id}/extract")
async def get_ledger_extract(
    supplier_id: int,
    empresa_id: Optional[int] = Query(None),
    start_date: Optional[str] = Query(None, description="YYYY-MM-DD"),
    end_date: Optional[str] = Query(None, description="YYYY-MM-DD"),
    svc: AccountingMatchService = Depends(get_accounting_service),
):
    """Extrato contabilístico por período com Saldo Inicial (tipo_doc FT/NC/ND/RE, D/C)."""
    try:
        return svc.get_ledger_extract_by_period(supplier_id, empresa_id, start_date, end_date)
    finally:
        svc.close()


@router.get("/ledger/{supplier_id}", response_model=LedgerResponse)
async def get_ledger(
    supplier_id: int,
    empresa_id: Optional[int] = Query(None),
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0),
    svc: AccountingMatchService = Depends(get_accounting_service),
):
    """Extrato da conta corrente de um fornecedor."""
    try:
        entries, total, saldo = svc.get_ledger(supplier_id, empresa_id, limit, offset)
        lines = [LedgerLine(**{k: str(v) if hasattr(v, 'isoformat') else v for k, v in e.items()}) for e in entries]
        return LedgerResponse(entries=lines, total=total, saldo_atual=saldo)
    finally:
        svc.close()


@router.post("/ledger/entry")
async def create_ledger_entry(
    body: LedgerEntryIn,
    svc: AccountingMatchService = Depends(get_accounting_service),
):
    """Cria um lançamento manual na conta corrente (Fatura, Pagamento, Nota de Crédito, Ajuste)."""
    from datetime import date as ddate
    try:
        dm = ddate.fromisoformat(body.data_movimento) if body.data_movimento else None
        result = svc.create_ledger_entry(
            empresa_id=body.empresa_id,
            supplier_id=body.supplier_id,
            tipo=body.tipo,
            valor_credito=body.valor_credito,
            valor_debito=body.valor_debito,
            documento_ref=body.documento_ref,
            purchase_order_id=body.purchase_order_id,
            notas=body.notas,
            data_movimento=dm,
        )
        return result
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        svc.close()


@router.get("/aging", response_model=List[AgingRow])
async def get_aging(
    empresa_id: Optional[int] = Query(None),
    svc: AccountingMatchService = Depends(get_accounting_service),
):
    """Aging report: dívidas a fornecedores divididas em A Vencer / < 30d / > 30d."""
    try:
        return svc.get_aging_report(empresa_id)
    finally:
        svc.close()


@router.get("/payments/antecipado", response_model=PaymentsAntecipadoResponse)
async def get_payments_antecipado(
    empresa_id: Optional[int] = Query(None),
    svc: AccountingMatchService = Depends(get_accounting_service),
):
    """Lista POs com prazo Antecipado ainda não pagas (para pagamento diário em bulk)."""
    try:
        items = svc.get_payments_antecipado(empresa_id)
        return PaymentsAntecipadoResponse(items=[PaymentSuggestionItem(**x) for x in items], total=len(items))
    finally:
        svc.close()


@router.get("/payments/sugestao", response_model=PaymentsSugestaoResponse)
async def get_payments_sugestao(
    empresa_id: Optional[int] = Query(None),
    data_inicio: Optional[str] = Query(None, description="YYYY-MM-DD início do intervalo de vencimento"),
    data_fim: Optional[str] = Query(None, description="YYYY-MM-DD fim do intervalo de vencimento"),
    svc: AccountingMatchService = Depends(get_accounting_service),
):
    """Sugestão de pagamento: POs com vencimento no intervalo [data_inicio, data_fim] ainda não pagas."""
    try:
        items, total_valor = svc.get_payments_sugestao(
            empresa_id=empresa_id,
            data_inicio=data_inicio or "",
            data_fim=data_fim or "",
        )
        return PaymentsSugestaoResponse(items=[PaymentSuggestionItem(**x) for x in items], total_valor=total_valor)
    finally:
        svc.close()


@router.post("/payments/confirmar", response_model=ConfirmPaymentResponse)
async def confirmar_pagamentos(
    body: ConfirmPaymentRequest,
    svc: AccountingMatchService = Depends(get_accounting_service),
):
    """Confirma pagamento no banco: cria lançamentos na conta corrente, marca POs como Paid e opcionalmente cria movimento de banco."""
    from datetime import date as ddate
    try:
        data_pag = None
        if body.data_pagamento:
            try:
                data_pag = ddate.fromisoformat(body.data_pagamento)
            except ValueError:
                raise HTTPException(status_code=400, detail="data_pagamento inválida. Use YYYY-MM-DD.")
        items_raw = [{"purchase_order_id": x.purchase_order_id, "valor": x.valor} for x in body.items]
        result = svc.confirmar_pagamentos(
            empresa_id=body.empresa_id,
            items=items_raw,
            data_pagamento=data_pag,
            criar_movimento_banco=body.criar_movimento_banco,
        )
        return ConfirmPaymentResponse(**result)
    finally:
        svc.close()


@router.get("/payments/historico", response_model=List[PaymentHistoricoItem])
async def get_payments_historico(
    empresa_id: Optional[int] = Query(None),
    data_inicio: Optional[str] = Query(None, description="YYYY-MM-DD"),
    data_fim: Optional[str] = Query(None, description="YYYY-MM-DD"),
    metodo: Optional[str] = Query(None, description="Cartao, Transferencia, DebitoDireto"),
    svc: AccountingMatchService = Depends(get_accounting_service),
):
    """Lista pagamentos já confirmados (por data e opcionalmente por método de pagamento do fornecedor)."""
    try:
        items = svc.get_payments_historico(
            empresa_id=empresa_id,
            data_inicio=data_inicio or "",
            data_fim=data_fim or "",
            metodo=metodo,
        )
        return [PaymentHistoricoItem(**x) for x in items]
    finally:
        svc.close()


@router.get("/payments/historico/export")
async def export_payments_historico_excel(
    empresa_id: Optional[int] = Query(None),
    data_inicio: Optional[str] = Query(None, description="YYYY-MM-DD"),
    data_fim: Optional[str] = Query(None, description="YYYY-MM-DD"),
    metodo: Optional[str] = Query(None),
    svc: AccountingMatchService = Depends(get_accounting_service),
):
    """Exporta o histórico de pagamentos (filtrado) para Excel."""
    try:
        items = svc.get_payments_historico(
            empresa_id=empresa_id,
            data_inicio=data_inicio or "",
            data_fim=data_fim or "",
            metodo=metodo,
        )
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        wb = Workbook()
        ws = wb.active
        ws.title = "Historico Pagamentos"
        headers = ["Data", "Empresa", "Fornecedor", "Método", "PO #", "Ref. doc", "Valor (€)", "Notas"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
            ws.cell(row=1, column=col).font = Font(bold=True)
        for row_idx, it in enumerate(items, 2):
            ws.cell(row=row_idx, column=1, value=it.get("data_movimento"))
            ws.cell(row=row_idx, column=2, value=it.get("empresa_nome") or "")
            ws.cell(row=row_idx, column=3, value=it.get("supplier_nome") or "")
            ws.cell(row=row_idx, column=4, value=it.get("metodo_pagamento") or "")
            ws.cell(row=row_idx, column=5, value=it.get("purchase_order_id"))
            ws.cell(row=row_idx, column=6, value=it.get("documento_ref") or "")
            ws.cell(row=row_idx, column=7, value=float(it.get("valor") or 0))
            ws.cell(row=row_idx, column=8, value=it.get("notas") or "")
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=historico_pagamentos.xlsx"},
        )
    finally:
        svc.close()


@router.get("/reconciliation/discrepancies", response_model=DiscrepanciesResponse)
async def get_discrepancies(
    empresa_id: Optional[int] = Query(None),
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0),
    svc: AccountingMatchService = Depends(get_accounting_service),
):
    """Lista divergências de valores entre PO, Fatura e Banco."""
    try:
        items, total = svc.get_discrepancies(empresa_id, limit, offset)
        rows = [DiscrepancyRow(**{k: str(v) if hasattr(v, 'isoformat') else v for k, v in r.items()}) for r in items]
        return DiscrepanciesResponse(items=rows, total=total)
    finally:
        svc.close()


@router.post("/reconciliation/match/{po_id}", response_model=MatchResult)
async def run_triple_match(
    po_id: int,
    svc: AccountingMatchService = Depends(get_accounting_service),
):
    """Executa Triple-Match para uma Purchase Order."""
    try:
        result = svc.run_triple_match(po_id)
        if "error" in result:
            raise HTTPException(status_code=404, detail=result["error"])
        return MatchResult(**result)
    finally:
        svc.close()


@router.get("/profitability", response_model=ProfitabilityResponse)
async def get_profitability(
    empresa_id: Optional[int] = Query(None),
    data_inicio: Optional[str] = Query(None, description="YYYY-MM-DD"),
    data_fim: Optional[str] = Query(None, description="YYYY-MM-DD"),
    svc: AccountingMatchService = Depends(get_accounting_service),
):
    """Lucro Real Final: GMV - Devoluções - Comissões - Custo Real - Portes."""
    try:
        return svc.get_profitability(empresa_id, data_inicio, data_fim)
    finally:
        svc.close()


@router.get("/cash-flow-forecast", response_model=List[CashFlowDay])
async def get_cash_flow_forecast(
    empresa_id: Optional[int] = Query(None),
    days: int = Query(30, ge=7, le=90, description="Horizonte em dias"),
    svc: AccountingMatchService = Depends(get_accounting_service),
):
    """Projeção de saídas de caixa por data de vencimento das POs não pagas."""
    try:
        return svc.get_cash_flow_forecast(empresa_id, days)
    finally:
        svc.close()


class CashFlowProjectionDay(BaseModel):
    data: str
    valor_dia: float
    saldo_acumulado: float


@router.get("/cash-flow-projection", response_model=List[CashFlowProjectionDay])
async def get_cash_flow_projection(
    empresa_id: Optional[int] = Query(None),
    days: int = Query(30, ge=7, le=90),
    initial_balance: float = Query(0.0, description="Saldo inicial (ex: saldo bancário atual)"),
    svc: AccountingMatchService = Depends(get_accounting_service),
):
    """v3.0: Projeção diária de saldo (entradas - saídas) para AreaChart."""
    try:
        return svc.get_cash_flow_projection(empresa_id=empresa_id, days=days, initial_balance=initial_balance)
    finally:
        svc.close()


class SupplierHealthRow(BaseModel):
    supplier_id: int
    supplier_nome: str
    health_score: float
    lead_time_days: Optional[float] = None
    return_rate_pct: float
    margin_alert_count: int
    avg_margin_pct: Optional[float] = None


class OssReportResponse(BaseModel):
    quarter: str
    by_country: Dict[str, Dict[str, float]]
    totals: Dict[str, float]


@router.get("/oss-report", response_model=OssReportResponse)
async def get_oss_report_endpoint(
    year: int = Query(..., ge=2020, le=2100),
    quarter: int = Query(..., ge=1, le=4),
    country: Optional[str] = Query(None),
    empresa_id: Optional[int] = Query(None),
):
    """Resumo OSS: impostos por trimestre e por país (IVA nacional vs destino UE)."""
    conn = get_db_connection()
    try:
        data = get_oss_report(conn, year=year, quarter=quarter, country=country, empresa_id=empresa_id)
        return OssReportResponse(**data)
    finally:
        conn.close()


@router.get("/supplier-health", response_model=List[SupplierHealthRow])
async def get_supplier_health(
    empresa_id: Optional[int] = Query(None),
):
    """v3.0: Ranking de fornecedores por health score e margem média (para Dashboard de Compras)."""
    svc = SupplierHealthService()
    try:
        return svc.get_all_suppliers_ranked(empresa_id=empresa_id)
    finally:
        svc.close()


# ── Módulo de Faturas ERP-Grade ───────────────────────────────────────────────


@router.get("/suppliers/{supplier_id}/open-pos")
async def get_supplier_open_pos(
    supplier_id: int,
    empresa_id: Optional[int] = Query(None),
):
    """Lista POs do fornecedor em aberto (Draft/Ordered/Paid) para seleção no modal de fatura."""
    svc = InvoiceService()
    try:
        return svc.get_open_pos_for_supplier(supplier_id=supplier_id, empresa_id=empresa_id)
    finally:
        svc.close()


@router.get("/invoices")
async def list_supplier_invoices(
    empresa_id: Optional[int] = Query(None),
    supplier_id: Optional[int] = Query(None),
    status: Optional[str] = Query(None),
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0),
):
    """Lista faturas de fornecedores com POs associadas."""
    svc = InvoiceService()
    try:
        items, total = svc.list_invoices(
            empresa_id=empresa_id, supplier_id=supplier_id,
            status=status, limit=limit, offset=offset,
        )
        return {"items": items, "total": total}
    finally:
        svc.close()


class CreateInvoiceBody(BaseModel):
    empresa_id: int
    supplier_id: int
    invoice_ref: str
    invoice_amount: float
    invoice_date: Optional[str] = None   # YYYY-MM-DD
    po_ids: List[int]
    notas: Optional[str] = None
    post_to_ledger: bool = True


@router.post("/invoices")
async def create_supplier_invoice(body: CreateInvoiceBody):
    """
    Regista fatura de fornecedor associada a uma ou mais POs.
    Cria lançamento único na Conta Corrente se post_to_ledger=true.
    """
    if not body.po_ids:
        from fastapi import HTTPException
        raise HTTPException(status_code=422, detail="po_ids não pode estar vazio")
    svc = InvoiceService()
    try:
        result = svc.create_invoice(
            empresa_id=body.empresa_id,
            supplier_id=body.supplier_id,
            invoice_ref=body.invoice_ref,
            invoice_amount=body.invoice_amount,
            invoice_date=body.invoice_date,
            po_ids=body.po_ids,
            notas=body.notas,
            post_to_ledger=body.post_to_ledger,
        )
        return {"success": True, **result}
    finally:
        svc.close()


@router.get("/invoices/{invoice_id}")
async def get_supplier_invoice(invoice_id: int):
    """Detalhe de uma fatura com lista de POs associadas."""
    svc = InvoiceService()
    try:
        inv = svc.get_invoice(invoice_id)
        if not inv:
            from fastapi import HTTPException
            raise HTTPException(status_code=404, detail="Fatura não encontrada")
        return inv
    finally:
        svc.close()
